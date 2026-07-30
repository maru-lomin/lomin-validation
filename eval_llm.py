"""LLM 기반 pred/gt 채점: detail.xlsx의 {필드명}_Pred / {필드명}_GT 쌍만 사용.

채점·출력 CSV에는 해당 Pred/GT 열만 사용합니다. 채점 후 detail.xlsx의 각 필드 _WER 우측에 _Score를 기록합니다.
열 구성: 파일명 열이 있으면 맨 앞에 한 번, 이후 각 필드마다
{필드}_Pred, {필드}_GT, {필드}_Score, {필드}_Reason, {필드}_Precision, {필드}_Recall 순입니다.
시트가 둘 이상이면 맨 앞에 eval_sheet, eval_excel_row 열을 추가합니다.
"""

from __future__ import annotations

import sys

# 무거운 import 전에 한 줄 (Python 프로세스만 떠도 바로 보임)
print("[eval_llm] Python 실행됨, 라이브러리 로딩 중…", file=sys.stderr, flush=True)

import argparse
import csv
import json
import os
import re
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from eval_llm_client import chat_completion, first_model_id_or_raise, get_model_catalog

print("[eval_llm] import 완료", file=sys.stderr, flush=True)


@dataclass(frozen=True)
class PredGtPair:
    field: str
    pred_col: int  # 1-based openpyxl column index
    gt_col: int


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    return s


def _find_pred_gt_pairs(headers: list[str]) -> list[PredGtPair]:
    name_to_idx: dict[str, int] = {}
    for i, h in enumerate(headers, start=1):
        if h is None:
            continue
        name = str(h).strip()
        if name:
            name_to_idx[name] = i
    pairs: list[PredGtPair] = []
    for name, idx in name_to_idx.items():
        if not name.endswith("_Pred"):
            continue
        base = name[: -len("_Pred")]
        gt_name = f"{base}_GT"
        if gt_name in name_to_idx:
            pairs.append(PredGtPair(field=base, pred_col=idx, gt_col=name_to_idx[gt_name]))
    pairs.sort(key=lambda p: p.pred_col)
    return pairs


def _filename_col_index(headers: list[str]) -> int | None:
    for i, h in enumerate(headers, start=1):
        if h is None:
            continue
        if str(h).strip() == "파일명":
            return i
    return None


def _header_cell_name(headers: list[str], col_idx: int) -> str:
    """CSV/딕셔너리 키용: 빈 헤더는 고유 플레이스홀더."""
    if col_idx < 0 or col_idx >= len(headers):
        return f"__col_{col_idx + 1}"
    h = headers[col_idx]
    hs = str(h).strip() if h is not None else ""
    return hs if hs else f"__unnamed_{col_idx + 1}"


def _build_wide_csv_fieldnames(headers: list[str], pairs: list[PredGtPair]) -> list[str]:
    """Pred/GT 및 LLM 채점 열만: [파일명]? + 필드별 Pred, GT, Score, Reason, Precision, Recall."""
    out: list[str] = []
    fn_col = _filename_col_index(headers)
    if fn_col is not None:
        out.append(_header_cell_name(headers, fn_col - 1))
    for p in pairs:
        out.extend(
            [
                f"{p.field}_Pred",
                f"{p.field}_GT",
                f"{p.field}_Score",
                f"{p.field}_Reason",
                f"{p.field}_Precision",
                f"{p.field}_Recall",
            ]
        )
    return out


def _ordered_union_fieldnames(lists: list[list[str]]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for cols in lists:
        for c in cols:
            if c not in seen:
                seen.add(c)
                out.append(c)
    return out


@dataclass
class SheetEvalBlock:
    title: str
    headers: list[str]
    pairs: list[PredGtPair]
    fn_col: int | None
    data_rows: list[tuple[int, tuple[Any, ...]]] = field(default_factory=list)


@dataclass(frozen=True)
class EvalTask:
    """채점 API 1건: 시트·엑셀 행·필드·pred/gt 및 전체 순서(order)."""

    order: int
    sheet: str
    excel_row: int
    filename: str
    field: str
    pred: str
    gt: str


@dataclass(frozen=True)
class EvalLlmRuntime:
    """태스크 실행에 공통으로 쓰는 설정(프롬프트·엔드포인트·로그 플래그)."""

    prompt_template: str
    base_url: str
    model: str
    timeout: float
    verbose: bool
    log_each_response: bool


def _wide_output_row(
    block: SheetEvalBlock,
    excel_row: int,
    row: tuple[Any, ...],
    cell_eval: dict[tuple[str, int, str], dict[str, str]],
    fieldnames: list[str],
    *,
    include_sheet_and_row: bool,
) -> dict[str, str]:
    """wide CSV 한 행: 파일명(있으면) + 각 필드 Pred/GT + 채점 열."""
    out: dict[str, str] = {k: "" for k in fieldnames}
    if include_sheet_and_row:
        out["eval_sheet"] = block.title
        out["eval_excel_row"] = str(excel_row)
    if block.fn_col is not None:
        fn_key = _header_cell_name(block.headers, block.fn_col - 1)
        if fn_key in out:
            val = (
                row[block.fn_col - 1]
                if block.fn_col <= len(row)
                else None
            )
            out[fn_key] = _cell_str(val)
    for p in block.pairs:
        pred_val = row[p.pred_col - 1] if p.pred_col <= len(row) else None
        gt_val = row[p.gt_col - 1] if p.gt_col <= len(row) else None
        pk, gk = f"{p.field}_Pred", f"{p.field}_GT"
        if pk in out:
            out[pk] = _cell_str(pred_val)
        if gk in out:
            out[gk] = _cell_str(gt_val)
        ev = cell_eval.get((block.title, excel_row, p.field), {})
        for suffix, ev_key in (
            ("_Score", "score"),
            ("_Reason", "reason"),
            ("_Precision", "precision"),
            ("_Recall", "recall"),
        ):
            col = f"{p.field}{suffix}"
            if col in out:
                out[col] = ev.get(ev_key, "")
    return out


def _extract_json_object(text: str) -> dict[str, Any] | None:
    text = text.strip()
    m = re.search(r"```(?:json)?\s*(\{.*?\})\s*```", text, re.DOTALL | re.IGNORECASE)
    if m:
        try:
            return json.loads(m.group(1))
        except json.JSONDecodeError:
            pass
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass
    start = text.find("{")
    if start < 0:
        return None
    depth = 0
    for i in range(start, len(text)):
        if text[i] == "{":
            depth += 1
        elif text[i] == "}":
            depth -= 1
            if depth == 0:
                try:
                    return json.loads(text[start : i + 1])
                except json.JSONDecodeError:
                    return None
    return None


def _snippet_for_log(text: str, max_chars: int = 4000) -> str:
    """터미널 로그용: 과도하게 긴 pred/gt 는 앞부분만."""
    if len(text) <= max_chars:
        return text
    return text[:max_chars] + f"\n… ({len(text)}자 중 앞 {max_chars}자만 표시)"


def _print_llm_completion_block(
    *,
    label: str,
    sheet: str,
    excel_row: int,
    filename: str,
    field: str,
    gt: str,
    pred: str,
    score: str,
    precision: str,
    recall: str,
    reason: str,
    raw: str,
    print_lock: threading.Lock | None,
    max_raw_chars: int = 12000,
    max_reason_chars: int = 800,
) -> None:
    """요청 1건 완료 시 gt/pred·파싱 결과·모델 원문을 터미널(stderr)에 출력."""
    tail = ""
    body = raw
    if len(body) > max_raw_chars:
        body = body[:max_raw_chars]
        tail = f"\n… (원문 {len(raw)}자 중 앞 {max_raw_chars}자만 표시)"
    reason_show = reason
    if len(reason_show) > max_reason_chars:
        reason_show = reason_show[:max_reason_chars] + f"… ({len(reason)}자 중 앞 {max_reason_chars}자)"
    lines = [
        "--------------------------------",
        f"{label} field={field!r}",
        f"filename={filename!r}",
        "--------------------------------",
        "[Ground Truth]",
        _snippet_for_log(gt),
        "--------------------------------",
        "[Prediction]",
        _snippet_for_log(pred),
        "--------------------------------",
        "[Model Response]",
        body + tail,
        "================================\n\n",
    ]
    text = "\n".join(lines) + "\n"

    def _emit() -> None:
        print(text, file=sys.stderr, end="", flush=True)

    if print_lock is not None:
        with print_lock:
            _emit()
    else:
        _emit()


def _normalize_scores(obj: dict[str, Any]) -> tuple[str, str, str, str]:
    reason = obj.get("reason", "")
    if not isinstance(reason, str):
        reason = str(reason)
    score = obj.get("score", "")
    precision = obj.get("precision", "")
    recall = obj.get("recall", "")
    out_score = str(score) if isinstance(score, (int, float, str)) else str(score)
    out_prec = str(precision) if isinstance(precision, (int, float, str)) else str(precision)
    out_rec = str(recall) if isinstance(recall, (int, float, str)) else str(recall)
    return out_score, reason, out_prec, out_rec


def _progress_step(total: int) -> int:
    return max(1, total // 25) if total else 1


def _maybe_report_batch_progress(
    done: int,
    total: int,
    *,
    log_each_response: bool,
) -> None:
    """--no-echo-llm 시 대량 호출에서 간헐적 진행만 stderr에 출력."""
    if log_each_response or not total:
        return
    step = _progress_step(total)
    if done == total or done % step == 0:
        print(f"[eval_llm] 진행 {done}/{total}", file=sys.stderr, flush=True)


def load_plans_and_tasks(pred_gt_file: Path) -> tuple[list[SheetEvalBlock], list[EvalTask]]:
    """워크북에서 Pred/GT 쌍이 있는 시트만 읽어 블록·평가 태스크 목록을 만든다."""
    plans: list[SheetEvalBlock] = []
    tasks: list[EvalTask] = []
    order = 0
    wb = load_workbook(pred_gt_file, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not first:
                continue
            headers = [str(h) if h is not None else "" for h in first]
            pairs = _find_pred_gt_pairs(headers)
            if not pairs:
                continue
            fn_col = _filename_col_index(headers)
            block = SheetEvalBlock(title=ws.title, headers=headers, pairs=pairs, fn_col=fn_col)
            for excel_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if row is None or all(c is None or str(c).strip() == "" for c in row):
                    continue
                filename = ""
                if fn_col is not None and fn_col <= len(row):
                    filename = _cell_str(row[fn_col - 1])
                block.data_rows.append((excel_row, row))
                for pair in pairs:
                    pred = _cell_str(
                        row[pair.pred_col - 1] if pair.pred_col <= len(row) else None
                    )
                    gt = _cell_str(row[pair.gt_col - 1] if pair.gt_col <= len(row) else None)
                    tasks.append(
                        EvalTask(
                            order=order,
                            sheet=ws.title,
                            excel_row=excel_row,
                            filename=filename,
                            field=pair.field,
                            pred=pred,
                            gt=gt,
                        )
                    )
                    order += 1
            plans.append(block)
    finally:
        wb.close()
    return plans, tasks


def build_cell_eval(
    tasks: list[EvalTask], results: list[dict[str, str]]
) -> dict[tuple[str, int, str], dict[str, str]]:
    cell_eval: dict[tuple[str, int, str], dict[str, str]] = {}
    for task, r in zip(tasks, results, strict=True):
        cell_eval[(task.sheet, task.excel_row, task.field)] = {
            "score": r["score"],
            "reason": r["reason"],
            "precision": r["precision"],
            "recall": r["recall"],
        }
    return cell_eval


def execute_eval_tasks(
    tasks: list[EvalTask],
    *,
    runtime: EvalLlmRuntime,
    max_workers: int,
) -> list[dict[str, str]]:
    """태스크 순서대로 API를 호출하고, 결과 dict 목록을 tasks와 동일 순서로 반환."""
    total = len(tasks)
    print_lock: threading.Lock | None = threading.Lock() if max_workers > 1 else None
    if max_workers <= 1:
        results: list[dict[str, str]] = []
        for task in tasks:
            row_out, _ = _run_one(task, total, runtime, print_lock=None)
            results.append(row_out)
            _maybe_report_batch_progress(
                len(results), total, log_each_response=runtime.log_each_response
            )
        return results

    indexed: dict[int, dict[str, str]] = {}
    done = 0
    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        futs = {
            ex.submit(_run_one, task, total, runtime, print_lock): task.order for task in tasks
        }
        for fut in as_completed(futs):
            o = futs[fut]
            row_out, _ = fut.result()
            indexed[o] = row_out
            done += 1
            _maybe_report_batch_progress(done, total, log_each_response=runtime.log_each_response)
    return [indexed[i] for i in range(len(tasks))]


def build_wide_csv_rows(
    plans: list[SheetEvalBlock],
    cell_eval: dict[tuple[str, int, str], dict[str, str]],
) -> tuple[list[str], list[dict[str, str]]]:
    include_sheet_and_row = len(plans) > 1
    wide_per_plan = [_build_wide_csv_fieldnames(b.headers, b.pairs) for b in plans]
    if include_sheet_and_row:
        csv_fieldnames = ["eval_sheet", "eval_excel_row"] + _ordered_union_fieldnames(wide_per_plan)
    else:
        csv_fieldnames = _ordered_union_fieldnames(wide_per_plan) if plans else []

    wide_rows: list[dict[str, str]] = []
    for block in plans:
        for excel_row, row in block.data_rows:
            wide_rows.append(
                _wide_output_row(
                    block,
                    excel_row,
                    row,
                    cell_eval,
                    csv_fieldnames,
                    include_sheet_and_row=include_sheet_and_row,
                )
            )
    return csv_fieldnames, wide_rows


def write_wide_csv_file(
    output_file: Path, fieldnames: list[str], wide_rows: list[dict[str, str]]
) -> None:
    output_file.parent.mkdir(parents=True, exist_ok=True)
    with output_file.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        w.writeheader()
        w.writerows(wide_rows)



def scores_from_wide_rows(wide_rows: list[dict[str, str]]) -> dict[tuple[str, str], str]:
    """(파일명, 필드명) → Score 문자열."""
    out: dict[tuple[str, str], str] = {}
    for row in wide_rows:
        fn = (row.get("파일명") or "").strip()
        if not fn:
            continue
        for k, v in row.items():
            if k.endswith("_Score"):
                out[(fn, k[: -len("_Score")])] = v if v is not None else ""
    return out


def _header_names(ws) -> list[str]:
    return [
        str(c.value).strip() if c.value is not None else ""
        for c in next(ws.iter_rows(min_row=1, max_row=1))
    ]


def _rebuild_wide_headers_with_score(headers: list[str]) -> list[str]:
    """각 필드의 _WER 바로 뒤에 _Score가 오도록 헤더를 재구성."""
    out: list[str] = []
    i = 0
    while i < len(headers):
        h = headers[i]
        if h.endswith("_Score"):
            i += 1
            continue
        out.append(h)
        if h.endswith("_WER"):
            base = h[: -len("_WER")]
            score_h = f"{base}_Score"
            # 기존에 바로 다음이 Score면 소비만 하고, 아니면 삽입
            if i + 1 < len(headers) and headers[i + 1] == score_h:
                out.append(score_h)
                i += 2
                continue
            out.append(score_h)
        i += 1
    return out


def write_llm_scores_to_detail_xlsx(
    detail_path: Path,
    scores: dict[tuple[str, str], str],
) -> None:
    """detail.xlsx의 파일별·파일·class별 시트에 Score를 WER 우측에 기록."""
    from openpyxl.styles import Font

    wb = load_workbook(detail_path)
    header_font = Font(bold=True)

    # --- 파일별 ---
    if "파일별" in wb.sheetnames:
        ws = wb["파일별"]
        old_headers = _header_names(ws)
        new_headers = _rebuild_wide_headers_with_score(old_headers)
        old_idx = {h: i for i, h in enumerate(old_headers) if h}
        fn_i = old_idx.get("파일명")

        data_rows: list[list] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            vals = list(row)
            # pad
            if len(vals) < len(old_headers):
                vals.extend([None] * (len(old_headers) - len(vals)))
            new_row: list = []
            filename = ""
            if fn_i is not None and fn_i < len(vals) and vals[fn_i] is not None:
                filename = str(vals[fn_i]).strip()
            for h in new_headers:
                if h.endswith("_Score"):
                    field = h[: -len("_Score")]
                    # 기존 Score 열이 있으면 우선 복사, scores 맵이 있으면 덮어씀
                    if h in old_idx and old_idx[h] < len(vals):
                        new_row.append(vals[old_idx[h]])
                    else:
                        new_row.append(None)
                    if filename and (filename, field) in scores:
                        new_row[-1] = scores[(filename, field)]
                elif h in old_idx:
                    new_row.append(vals[old_idx[h]])
                else:
                    new_row.append(None)
            data_rows.append(new_row)

        # 시트 재작성
        ws.delete_rows(1, ws.max_row or 1)
        ws.append(new_headers)
        for c in range(1, len(new_headers) + 1):
            ws.cell(row=1, column=c).font = header_font
        for r in data_rows:
            ws.append(r)
        ws.freeze_panes = "B2"

    # --- 파일·class별 ---
    long_name = "파일·class별"
    if long_name in wb.sheetnames:
        ws = wb[long_name]
        headers = _header_names(ws)
        # WER 다음에 Score 보장
        if "WER" in headers:
            wer_i = headers.index("WER")
            if "Score" in headers:
                # Score를 WER 바로 뒤로 이동
                score_i = headers.index("Score")
                if score_i != wer_i + 1:
                    headers.pop(score_i)
                    if score_i < wer_i:
                        wer_i = headers.index("WER")
                    headers.insert(wer_i + 1, "Score")
            else:
                headers.insert(wer_i + 1, "Score")

        name_i = headers.index("파일명") if "파일명" in headers else None
        cls_i = headers.index("KV class") if "KV class" in headers else None
        score_i = headers.index("Score") if "Score" in headers else None

        # 기존 열 맵 (Score 삽입 전 원본)
        old_headers = _header_names(ws)
        old_idx = {h: i for i, h in enumerate(old_headers) if h}

        data_rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            vals = list(row)
            if len(vals) < len(old_headers):
                vals.extend([None] * (len(old_headers) - len(vals)))
            new_row = []
            for h in headers:
                if h == "Score":
                    fn = ""
                    cls = ""
                    if name_i is not None and "파일명" in old_idx:
                        v = vals[old_idx["파일명"]]
                        fn = str(v).strip() if v is not None else ""
                    if cls_i is not None and "KV class" in old_idx:
                        v = vals[old_idx["KV class"]]
                        cls = str(v).strip() if v is not None else ""
                    if "Score" in old_idx:
                        new_row.append(vals[old_idx["Score"]])
                    else:
                        new_row.append(None)
                    if fn and cls and (fn, cls) in scores:
                        new_row[-1] = scores[(fn, cls)]
                elif h in old_idx:
                    new_row.append(vals[old_idx[h]])
                else:
                    new_row.append(None)
            data_rows.append(new_row)

        ws.delete_rows(1, ws.max_row or 1)
        ws.append(headers)
        for c in range(1, len(headers) + 1):
            ws.cell(row=1, column=c).font = header_font
        for r in data_rows:
            ws.append(r)
        ws.freeze_panes = "A2"

    wb.save(detail_path)


def _run_one(
    task: EvalTask,
    total_tasks: int,
    runtime: EvalLlmRuntime,
    print_lock: threading.Lock | None,
) -> tuple[dict[str, str], str]:
    raw = ""
    if task.pred == "" and task.gt == "":
        if runtime.verbose:
            print(
                f"[eval_llm] 스킵 (pred·gt 모두 빈 문자열) ({task.order + 1}/{total_tasks}) "
                f"sheet={task.sheet!r} row={task.excel_row} field={task.field!r}",
                file=sys.stderr,
                flush=True,
            )
        out = {
            "sheet": task.sheet,
            "row": str(task.excel_row),
            "filename": task.filename,
            "field": task.field,
            "score": "",
            "reason": "",
            "precision": "",
            "recall": "",
        }
        return out, raw

    user = runtime.prompt_template.format(
        key_description=task.field, gt=task.gt, pred=task.pred
    )
    try:
        if runtime.verbose:
            print(
                f"[eval_llm] POST chat/completions ({task.order + 1}/{total_tasks}) "
                f"sheet={task.sheet!r} row={task.excel_row} field={task.field!r}",
                file=sys.stderr,
                flush=True,
            )
        raw = chat_completion(
            base_url=runtime.base_url,
            model=runtime.model,
            user_content=user,
            timeout=runtime.timeout,
        )
        obj = _extract_json_object(raw)
        if not obj:
            score, reason, prec, rec = (
                "",
                f"JSON 파싱 실패. 원문: {raw[:500]}",
                "",
                "",
            )
        else:
            score, reason, prec, rec = _normalize_scores(obj)
    except Exception as e:
        score, reason, prec, rec = ("", f"API 오류: {e}", "", "")
        if not raw:
            raw = f"(응답 본문 없음: {e})"
    out = {
        "sheet": task.sheet,
        "row": str(task.excel_row),
        "filename": task.filename,
        "field": task.field,
        "score": score,
        "reason": reason,
        "precision": prec,
        "recall": rec,
    }
    if runtime.log_each_response:
        _print_llm_completion_block(
            label=f"{task.order + 1}/{total_tasks}",
            sheet=task.sheet,
            excel_row=task.excel_row,
            filename=task.filename,
            field=task.field,
            gt=task.gt,
            pred=task.pred,
            score=score,
            precision=prec,
            recall=rec,
            reason=reason,
            raw=raw,
            print_lock=print_lock,
        )
    return out, raw


def run_probe(*, base_url: str, model: str | None, timeout: float) -> int:
    base = base_url.rstrip("/")
    murl = base + "/v1/models"
    print(f"[eval_llm:probe] GET {murl}", file=sys.stderr, flush=True)
    mids, data = get_model_catalog(base_url, timeout=timeout)
    if not mids:
        print(f"[eval_llm:probe] 오류: /v1/models 에 id 없음: {data!r}", file=sys.stderr)
        return 1
    raw_m = (model or "").strip()
    if not raw_m or raw_m.lower() == "auto":
        resolved = mids[0]
        tag = "--model auto" if raw_m.lower() == "auto" else "model 미지정"
        print(f"[eval_llm:probe] {tag} → 첫 모델 id: {resolved!r}", file=sys.stderr, flush=True)
    else:
        resolved = raw_m
        if resolved not in mids:
            print(
                f"[eval_llm:probe] 경고: 지정 모델이 목록에 없을 수 있음. 목록: {mids[:5]}...",
                file=sys.stderr,
                flush=True,
            )
        print(f"[eval_llm:probe] 지정 모델: {resolved!r}", file=sys.stderr, flush=True)
    curl = base + "/v1/chat/completions"
    print(f"[eval_llm:probe] POST {curl}", file=sys.stderr, flush=True)
    raw = chat_completion(
        base_url=base_url,
        model=resolved,
        user_content="ping",
        timeout=timeout,
        max_tokens=16,
    )
    print(f"[eval_llm:probe] 응답 앞부분: {raw[:120]!r}", file=sys.stderr, flush=True)
    print(
        "[eval_llm:probe] 완료. `docker logs -f eval-llm` 에 "
        '`POST /v1/chat/completions` 한 줄이 추가되었는지 확인하세요.',
        file=sys.stderr,
        flush=True,
    )
    return 0


@dataclass
class FieldAccuracyStats:
    field: str
    total: int = 0
    excluded: int = 0
    correct: int = 0
    wrong: int = 0

    @property
    def scored(self) -> int:
        return self.correct + self.wrong

    @property
    def accuracy(self) -> float | None:
        if self.scored == 0:
            return None
        return self.correct / self.scored


@dataclass
class AccuracySummary:
    result_csv: Path
    threshold: int
    row_count: int
    total: int
    excluded: int
    correct: int
    wrong: int
    per_field: list[FieldAccuracyStats]

    @property
    def scored(self) -> int:
        return self.correct + self.wrong

    @property
    def accuracy(self) -> float | None:
        if self.scored == 0:
            return None
        return self.correct / self.scored


def _parse_score_value(raw: str) -> int | None:
    s = raw.strip()
    if not s:
        return None
    try:
        return int(float(s))
    except ValueError:
        return None


def _is_correct_score(score: int, threshold: int) -> bool:
    return score >= threshold


def summarize_accuracy_from_csv(result_csv: Path, *, threshold: int) -> AccuracySummary:
    """eval_llm.csv wide 포맷에서 필드별·전체 정답률을 집계한다."""
    if not result_csv.is_file():
        raise FileNotFoundError(f"결과 CSV 없음: {result_csv}")

    with result_csv.open(encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames:
            raise ValueError(f"CSV 헤더 없음: {result_csv}")
        score_fields = [c for c in reader.fieldnames if c.endswith("_Score")]
        kv_fields = [c[: -len("_Score")] for c in score_fields]
        per_field_map: dict[str, FieldAccuracyStats] = {
            field: FieldAccuracyStats(field=field) for field in kv_fields
        }
        total = excluded = correct = wrong = 0
        row_count = 0
        for row in reader:
            row_count += 1
            for field in kv_fields:
                total += 1
                fs = per_field_map[field]
                fs.total += 1
                score = _parse_score_value(row.get(f"{field}_Score", "") or "")
                if score is None:
                    excluded += 1
                    fs.excluded += 1
                    continue
                if _is_correct_score(score, threshold):
                    correct += 1
                    fs.correct += 1
                else:
                    wrong += 1
                    fs.wrong += 1

    per_field = [per_field_map[f] for f in kv_fields]
    return AccuracySummary(
        result_csv=result_csv,
        threshold=threshold,
        row_count=row_count,
        total=total,
        excluded=excluded,
        correct=correct,
        wrong=wrong,
        per_field=per_field,
    )


def _fmt_pct(value: float | None) -> str:
    if value is None:
        return "N/A"
    return f"{value * 100:.2f}%"


def print_accuracy_summary(summary: AccuracySummary) -> None:
    lines = [
        "",
        "=" * 72,
        "[eval_llm] 정확도 요약",
        f"  CSV: {summary.result_csv}",
        f"  정답 기준: score >= {summary.threshold}",
        f"  문서(행) 수: {summary.row_count}",
        "-" * 72,
        f"  전체 셀 수: {summary.total}",
        f"  채점 제외: {summary.excluded} (pred·gt 모두 빈 값 등 score 없음)",
        f"  채점 대상: {summary.scored}",
        f"  맞음 (score>={summary.threshold}): {summary.correct}",
        f"  틀림 (score<{summary.threshold}): {summary.wrong}",
        f"  정확도: {_fmt_pct(summary.accuracy)} ({summary.correct}/{summary.scored})",
        "-" * 72,
        "  KV 필드별 정확도:",
    ]
    name_w = max((len(s.field) for s in summary.per_field), default=2)
    for fs in summary.per_field:
        lines.append(
            f"    {fs.field:<{name_w}}  "
            f"정확도 {_fmt_pct(fs.accuracy):>7}  "
            f"맞음 {fs.correct:>4}  틀림 {fs.wrong:>4}  "
            f"제외 {fs.excluded:>4}  전체 {fs.total:>4}"
        )
    lines.append("=" * 72)
    lines.append("")
    print("\n".join(lines), flush=True)


def run_summary_accuracy(result_csv: Path, *, threshold: int) -> int:
    try:
        summary = summarize_accuracy_from_csv(result_csv, threshold=threshold)
    except (FileNotFoundError, ValueError) as e:
        print(f"[eval_llm] 오류: {e}", file=sys.stderr)
        return 1
    print_accuracy_summary(summary)
    return 0


def parse_cli_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="LLM으로 detail.xlsx의 *_Pred / *_GT 쌍 채점")
    p.add_argument(
        "--summary_accuracy",
        action="store_true",
        help="--result_csv 의 eval_llm.csv 를 읽어 정확도 요약만 출력",
    )
    p.add_argument(
        "--result_csv",
        type=Path,
        default=None,
        help="--summary_accuracy 시 집계할 wide CSV 경로",
    )
    p.add_argument(
        "--threshold",
        type=int,
        default=10,
        help="--summary_accuracy 시 정답 판정 기준 (score >= threshold 이면 정답, 기본 10)",
    )
    p.add_argument("--pred_gt_file", type=Path, default=None)
    p.add_argument("--eval_prompt_file", type=Path, default=None)
    p.add_argument("--output_file", type=Path, default=None)
    p.add_argument(
        "--vllm_url",
        default=os.environ.get("VLLM_URL", "http://localhost:8000"),
        help="vLLM OpenAI 호환 베이스 URL (예: http://localhost:8002)",
    )
    p.add_argument(
        "--model",
        default=os.environ.get("VLLM_MODEL", ""),
        help='chat completions model id (/v1/models 의 id). 비우면 오류. "auto"면 첫 모델 사용.',
    )
    p.add_argument("--timeout", type=float, default=120.0)
    p.add_argument("--max-workers", type=int, default=4)
    p.add_argument(
        "--probe",
        action="store_true",
        help="엑셀 없이 /v1/models + 최소 chat 한 번만 호출 (연결·로그 확인용)",
    )
    p.add_argument(
        "--verbose",
        action="store_true",
        help="각 chat/completions 호출 직전에 stderr에 한 줄씩 출력",
    )
    p.add_argument(
        "--no-echo-llm",
        action="store_true",
        help="요청 완료마다 모델 원문 stderr 출력 끔 (로그만 간헐적 진행 표시)",
    )
    return p.parse_args()


def main() -> int:
    args = parse_cli_args()
    print("[eval_llm] 인자 파싱 완료, 실행 중…", file=sys.stderr, flush=True)

    if args.summary_accuracy:
        if args.result_csv is None:
            print("오류: --summary_accuracy 는 --result_csv 가 필요합니다.", file=sys.stderr)
            return 2
        return run_summary_accuracy(args.result_csv, threshold=args.threshold)

    if args.probe:
        return run_probe(base_url=args.vllm_url, model=args.model, timeout=args.timeout)

    if args.pred_gt_file is None or args.eval_prompt_file is None or args.output_file is None:
        print(
            "오류: --probe 가 아니면 --pred_gt_file, --eval_prompt_file, --output_file 가 필요합니다.",
            file=sys.stderr,
        )
        return 2

    model = (args.model or "").strip()
    if model.lower() == "auto":
        model = first_model_id_or_raise(args.vllm_url, timeout=args.timeout)
        print(f"[eval_llm] --model auto → 서버 모델 id: {model!r}", file=sys.stderr, flush=True)
    elif not model:
        print(
            "오류: --model 또는 환경변수 VLLM_MODEL을 지정하세요. "
            '(예: `curl -s $VLLM_URL/v1/models` 의 data[0].id 또는 --model auto)',
            file=sys.stderr,
        )
        return 2

    prompt_template = args.eval_prompt_file.read_text(encoding="utf-8")

    print(f"[eval_llm] 엑셀 읽는 중: {args.pred_gt_file}", file=sys.stderr, flush=True)
    plans, tasks = load_plans_and_tasks(args.pred_gt_file)

    log_each_response = not args.no_echo_llm
    endpoint = args.vllm_url.rstrip("/") + "/v1/chat/completions"
    echo_line = (
        "요청 1건 완료마다 모델 응답 원문을 stderr에 출력합니다. 끄려면 --no-echo-llm"
        if log_each_response
        else "모델 원문 출력 끔 (--no-echo-llm); 진행만 간헐적으로 stderr 출력"
    )
    print(
        f"[eval_llm] vLLM: {endpoint}\n"
        f"[eval_llm] model: {model!r}\n"
        f"[eval_llm] 채점 API 호출 예정: {len(tasks)}건. {echo_line}\n"
        f"[eval_llm] (--verbose: 각 요청 직전 한 줄)",
        file=sys.stderr,
        flush=True,
    )
    if not tasks:
        print(
            "[eval_llm] 경고: 호출 0건 → 서버 로그에 변화 없음. "
            "시트에 `필드_Pred` / `필드_GT` 헤더 쌍이 있는지, 데이터 행이 비어 있지 않은지 확인하세요.",
            file=sys.stderr,
            flush=True,
        )

    runtime = EvalLlmRuntime(
        prompt_template=prompt_template,
        base_url=args.vllm_url,
        model=model,
        timeout=args.timeout,
        verbose=args.verbose,
        log_each_response=log_each_response,
    )
    results = execute_eval_tasks(tasks, runtime=runtime, max_workers=args.max_workers)
    cell_eval = build_cell_eval(tasks, results)
    csv_fieldnames, wide_rows = build_wide_csv_rows(plans, cell_eval)
    write_wide_csv_file(args.output_file, csv_fieldnames, wide_rows)

    scores = scores_from_wide_rows(wide_rows)
    write_llm_scores_to_detail_xlsx(args.pred_gt_file, scores)
    print(
        f"[eval_llm] detail.xlsx Score 반영: {args.pred_gt_file} "
        f"({len(scores)}개 필드 점수)",
        file=sys.stderr,
        flush=True,
    )

    msg = f"저장: {args.output_file} ({len(wide_rows)}행, wide)"
    print(msg, flush=True)
    print(f"[eval_llm] {msg}", file=sys.stderr, flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
