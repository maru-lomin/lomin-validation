#!/usr/bin/env python3
"""
summary.txt + detail.xlsx → report.xlsx
  Sheet1: 필드별 F1 Score 요약
  Sheet2: 파일별 PRED, GT, F1 컬럼
"""
import argparse
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── 스타일 헬퍼 ──────────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT  = Font(color="FFFFFF", bold=True, size=11)
SUBHDR_FILL  = PatternFill("solid", fgColor="2E75B6")
SUBHDR_FONT  = Font(color="FFFFFF", bold=True)
TOTAL_FILL   = PatternFill("solid", fgColor="D6E4F0")
TOTAL_FONT   = Font(bold=True)
ALT_FILL     = PatternFill("solid", fgColor="EBF3FB")
CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT_WRAP    = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
LEFT         = Alignment(horizontal="left",   vertical="center")

def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def apply_header(ws, row, col, value, fill=HEADER_FILL, font=HEADER_FONT):
    c = ws.cell(row=row, column=col, value=value)
    c.fill  = fill
    c.font  = font
    c.alignment = CENTER
    c.border = thin_border()
    return c

def apply_cell(ws, row, col, value, alignment=CENTER, fill=None, font=None, number_format=None):
    c = ws.cell(row=row, column=col, value=value)
    c.alignment = alignment
    c.border = thin_border()
    if fill:   c.fill   = fill
    if font:   c.font   = font
    if number_format: c.number_format = number_format
    return c


# ── summary.txt 파싱 ─────────────────────────────────────────────────────────
def parse_summary(txt_path: Path):
    """
    Returns:
        overall: dict with Macro_F1, Micro_F1, ...
        by_class: list of dict {class, Macro_CER, Micro_CER, Macro_WER, Micro_WER, Macro_F1, Micro_F1, n}
    """
    text = txt_path.read_text(encoding="utf-8")

    # 전체 Macro F1
    m = re.search(r"Macro F1:\s*([\d.]+)", text)
    overall_macro_f1 = float(m.group(1)) if m else None

    # class별 섹션
    # 헤더: class      Macro_CER   Micro_CER   Macro_WER   Micro_WER    Macro_F1    Micro_F1     n
    by_class = []
    in_class_section = False
    for line in text.splitlines():
        if "전체 파일에 대한 class별 집계" in line:
            in_class_section = True
            continue
        if "전체 class에 대한 파일별 집계" in line:
            break
        if not in_class_section:
            continue
        line = line.strip()
        if not line or line.startswith("class") or line.startswith("---"):
            continue
        parts = line.split()
        if len(parts) < 8:
            continue
        try:
            row = {
                "class":     parts[0],
                "Macro_CER": float(parts[1]),
                "Micro_CER": float(parts[2]),
                "Macro_WER": float(parts[3]),
                "Micro_WER": float(parts[4]),
                "Macro_F1":  float(parts[5]),
                "Micro_F1":  float(parts[6]),
                "n":         int(parts[7]),
            }
            by_class.append(row)
        except (ValueError, IndexError):
            continue

    return overall_macro_f1, by_class


# ── detail.xlsx 파싱 ─────────────────────────────────────────────────────────
def parse_detail(xlsx_path: Path):
    """
    '파일별' 시트에서 각 파일, 각 필드의 Pred / GT / Word_F1(≈Macro_F1) 추출
    Returns:
        fields: list of field names (순서 유지)
        rows:   list of {filename, field -> {pred, gt, f1}}
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["파일별"]

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    # 필드 목록 추출 (헤더에서 _Pred 기준)
    fields = []
    field_idx = {}  # field -> {pred, gt, f1}
    for i, h in enumerate(headers):
        if h and h.endswith("_Pred"):
            field = h[:-5]  # remove _Pred
            fields.append(field)
            # find GT, Word_F1 indices
            gt_col  = headers.index(f"{field}_GT")   if f"{field}_GT"       in headers else None
            f1_col  = headers.index(f"{field}_Word_F1") if f"{field}_Word_F1" in headers else None
            field_idx[field] = {"pred": i, "gt": gt_col, "f1": f1_col}

    rows = []
    for ws_row in ws.iter_rows(min_row=2, values_only=True):
        filename = ws_row[0]
        if not filename:
            continue
        entry = {"filename": filename}
        for field in fields:
            idx = field_idx[field]
            pred = ws_row[idx["pred"]] if idx["pred"] is not None else None
            gt   = ws_row[idx["gt"]]   if idx["gt"]   is not None else None
            f1   = ws_row[idx["f1"]]   if idx["f1"]   is not None else None
            entry[field] = {"pred": pred, "gt": gt, "f1": f1}
        rows.append(entry)

    wb.close()
    return fields, rows


# ── Sheet1: 필드별 F1 요약 ───────────────────────────────────────────────────
def build_sheet1(wb, overall_macro_f1, by_class):
    ws = wb.create_sheet("Summary")

    # 제목 행
    ws.merge_cells("A1:C1")
    c = ws["A1"]
    c.value = "필드별 F1 Score 요약"
    c.fill  = HEADER_FILL
    c.font  = Font(color="FFFFFF", bold=True, size=13)
    c.alignment = CENTER
    ws.row_dimensions[1].height = 28

    # 헤더
    for col, label in enumerate(["필드 (Class)", "n (샘플수)", "F1"], start=1):
        apply_header(ws, 2, col, label)
    ws.row_dimensions[2].height = 22

    # 데이터
    for i, row in enumerate(by_class, start=3):
        fill = ALT_FILL if i % 2 == 0 else None
        apply_cell(ws, i, 1, row["class"], alignment=LEFT, fill=fill)
        apply_cell(ws, i, 2, row["n"],     fill=fill)
        apply_cell(ws, i, 3, row["Macro_F1"], fill=fill, number_format="0.0000")

    # 전체 합계 행
    total_row = 3 + len(by_class)
    apply_cell(ws, total_row, 1, "전체 (Overall)", alignment=LEFT, fill=TOTAL_FILL, font=TOTAL_FONT)
    apply_cell(ws, total_row, 2, sum(r["n"] for r in by_class), fill=TOTAL_FILL, font=TOTAL_FONT)
    apply_cell(ws, total_row, 3, overall_macro_f1, fill=TOTAL_FILL, font=TOTAL_FONT, number_format="0.0000")

    # 열 너비
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 14

    ws.freeze_panes = "A3"
    return ws


# ── Sheet2: 파일별 PRED / GT / F1 ───────────────────────────────────────────
def build_sheet2(wb, fields, rows):
    ws = wb.create_sheet("Detail")

    # 행 1: 필드 그룹 헤더 (병합)
    ws.cell(row=1, column=1, value="파일명").fill  = HEADER_FILL
    ws.cell(row=1, column=1).font  = HEADER_FONT
    ws.cell(row=1, column=1).alignment = CENTER
    ws.cell(row=1, column=1).border = thin_border()
    ws.merge_cells(start_row=1, end_row=2, start_column=1, end_column=1)

    col = 2
    for field in fields:
        ws.merge_cells(start_row=1, end_row=1, start_column=col, end_column=col + 2)
        c = ws.cell(row=1, column=col, value=field)
        c.fill      = SUBHDR_FILL
        c.font      = SUBHDR_FONT
        c.alignment = CENTER
        c.border    = thin_border()
        # 하위 헤더: PRED / GT / F1
        for j, label in enumerate(["PRED", "GT", "F1"], start=0):
            apply_header(ws, 2, col + j, label, fill=SUBHDR_FILL, font=SUBHDR_FONT)
        col += 3

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 22

    # 데이터 행
    for i, row in enumerate(rows, start=3):
        fill = ALT_FILL if i % 2 == 0 else None
        apply_cell(ws, i, 1, row["filename"], alignment=LEFT, fill=fill)
        col = 2
        for field in fields:
            fdata = row.get(field, {})
            pred  = fdata.get("pred")
            gt    = fdata.get("gt")
            f1    = fdata.get("f1")
            apply_cell(ws, i, col,     pred, alignment=LEFT_WRAP, fill=fill)
            apply_cell(ws, i, col + 1, gt,   alignment=LEFT_WRAP, fill=fill)
            f1_c = apply_cell(ws, i, col + 2, f1, fill=fill, number_format="0.0000")
            col += 3

    # 열 너비
    ws.column_dimensions["A"].width = 55
    col = 2
    for _ in fields:
        ws.column_dimensions[get_column_letter(col)].width     = 30  # PRED
        ws.column_dimensions[get_column_letter(col + 1)].width = 30  # GT
        ws.column_dimensions[get_column_letter(col + 2)].width = 12  # F1
        col += 3

    ws.freeze_panes = "B3"
    return ws


# ── main ─────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Generate report.xlsx from detail.xlsx + summary.txt")
    parser.add_argument("--summary-txt",  required=True)
    parser.add_argument("--detail-xlsx",  required=True)
    parser.add_argument("--output-xlsx",  required=True)
    args = parser.parse_args()

    summary_path = Path(args.summary_txt)
    detail_path  = Path(args.detail_xlsx)
    output_path  = Path(args.output_xlsx)

    print(f"[1/3] Parsing summary: {summary_path}")
    overall_macro_f1, by_class = parse_summary(summary_path)
    print(f"      Overall F1 = {overall_macro_f1}, {len(by_class)} classes")

    print(f"[2/3] Parsing detail:  {detail_path}")
    fields, rows = parse_detail(detail_path)
    print(f"      {len(fields)} fields, {len(rows)} files")

    print(f"[3/3] Writing report:  {output_path}")
    wb = openpyxl.Workbook()
    # 기본 시트 제거
    wb.remove(wb.active)

    build_sheet1(wb, overall_macro_f1, by_class)
    build_sheet2(wb, fields, rows)

    wb.save(output_path)
    print(f"Done → {output_path}")


if __name__ == "__main__":
    main()
