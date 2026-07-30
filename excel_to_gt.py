from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from demo_to_gt import build_via_entry, resolve_pdf_path, via_result_key

# (answer_sheet 열 이름, VIA/evaluation class 이름)
EXCEL_KV_COLUMNS: list[tuple[str, str]] = [
    ("계약자", "계약자"),
    ("피보험자", "피보험자"),
    ("주소", "주소"),
    ("보험기간", "보험기간"),
    ("보험금지급기준", "보험금지급기준"),
    ("소급담보일", "소급담보일"),
    ("담보지역", "담보지역"),
    ("재판관할국", "재판관할국"),
    ("업종", "업종"),
    ("담보_내용", "\ub2f4\ubcf4\ub0b4\uc6a9"),
    ("보상한도_공제금", "보상한도_공제금"),
    ("보험증권명", "보험증권명"),
    ("약관", "약관"),
    ("임의_출재율", "임의출재율"),
    ("임의_수수료율", "임의수수료율"),
    ("참고", "참고"),
]

FILE_NAME_COLUMN = "file_name"
ZERO_BOX = [0.0, 0.0, 0.0, 0.0]


def _cell_value(raw: Any) -> str | None:
    if raw is None:
        return None
    s = str(raw).strip()
    return s if s else None


def row_to_kv(row: dict[str, Any]) -> dict[str, Any]:
    """엑셀 한 행 → inference_result.kv 형식 (bbox 없음)."""
    kv: dict[str, Any] = {}
    for col, class_name in EXCEL_KV_COLUMNS:
        value = _cell_value(row.get(col))
        if value is None:
            continue
        kv[class_name] = {
            "text": [""],
            "value": [value],
            "box": [ZERO_BOX],
        }
    return kv


def resolve_pdf_for_row(dataset_dir: Path, file_name: str) -> Path | None:
    path = dataset_dir / file_name
    if path.is_file():
        return path
    return resolve_pdf_path(dataset_dir, Path(file_name).stem)


def convert_answer_sheet(
    excel_path: Path,
    dataset_dir: Path,
    *,
    require_pdf: bool = True,
) -> tuple[dict[str, Any], list[str]]:
    if not excel_path.is_file():
        raise FileNotFoundError(f"엑셀 파일 없음: {excel_path}")
    if not dataset_dir.is_dir():
        raise FileNotFoundError(f"dataset 디렉터리 없음: {dataset_dir}")

    root: dict[str, Any] = {}
    warnings: list[str] = []

    wb = load_workbook(excel_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header_row = next(rows, None)
        if not header_row:
            warnings.append("시트에 데이터가 없습니다.")
            return root, warnings

        header = [str(h).strip() if h is not None else "" for h in header_row]
        if FILE_NAME_COLUMN not in header:
            raise ValueError(
                f"'{FILE_NAME_COLUMN}' 열이 없습니다. 헤더: {header[:10]}..."
            )
        file_name_idx = header.index(FILE_NAME_COLUMN)

        missing_cols = [c for c, _ in EXCEL_KV_COLUMNS if c not in header]
        if missing_cols:
            warnings.append(
                f"엑셀에 없는 KV 열(건너뜀): {', '.join(missing_cols)}"
            )

        for row_num, row in enumerate(rows, start=2):
            if not row:
                continue
            file_name = _cell_value(
                row[file_name_idx] if file_name_idx < len(row) else None
            )
            if file_name is None:
                continue

            pdf_path = resolve_pdf_for_row(dataset_dir, file_name)
            if pdf_path is None:
                if require_pdf:
                    warnings.append(
                        f"행 {row_num}: dataset에 PDF 없음 ({file_name}) — 건너뜀"
                    )
                    continue
                file_size = 0
            else:
                file_name = pdf_path.name
                file_size = pdf_path.stat().st_size

            row_dict = {
                header[i]: row[i] if i < len(row) else None
                for i in range(len(header))
                if header[i]
            }
            kv = row_to_kv(row_dict)
            if not kv:
                warnings.append(f"행 {row_num}: KV 값 없음 ({file_name}) — 건너뜀")
                continue

            entry = build_via_entry(file_name, file_size, kv)
            key = via_result_key(file_name, file_size)
            if key in root:
                warnings.append(
                    f"행 {row_num}: 중복 파일 ({file_name}) — 마지막 행으로 덮어씀"
                )
            root[key] = entry
    finally:
        wb.close()

    return root, warnings


def build_parser() -> argparse.ArgumentParser:
    base = Path(__file__).resolve().parent
    p = argparse.ArgumentParser(
        description="answer_sheet.xlsx를 채점용 VIA gt.json 형식으로 변환합니다."
    )
    p.add_argument(
        "--excel",
        type=Path,
        default=base / "convert_sample" / "answer_sheet.xlsx",
        help="입력 엑셀 경로",
    )
    p.add_argument(
        "--dataset-dir",
        type=Path,
        default=base / "convert_sample" / "dataset",
        help="원본 PDF 디렉터리 (filename·size 매핑용)",
    )
    p.add_argument(
        "--output",
        type=Path,
        default=None,
        help="출력 gt.json 경로 (기본: excel 파일과 같은 디렉터리/gt.json)",
    )
    p.add_argument(
        "--include-missing-pdf",
        action="store_true",
        help="dataset에 PDF가 없어도 행을 포함 (size=0)",
    )
    return p


def main() -> None:
    args = build_parser().parse_args()
    excel_path = args.excel.resolve()
    dataset_dir = args.dataset_dir.resolve()
    output = args.output
    if output is None:
        output = excel_path.parent / "gt.json"
    else:
        output = output.resolve()

    root, warnings = convert_answer_sheet(
        excel_path,
        dataset_dir,
        require_pdf=not args.include_missing_pdf,
    )

    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", encoding="utf-8") as f:
        json.dump(root, f, ensure_ascii=False, indent=4)

    print(f"[완료] {len(root)}개 문서 → {output}")
    _print_warnings(warnings)


def _print_warnings(warnings: list[str], preview: int = 5) -> None:
    if not warnings:
        return
    from collections import Counter

    kinds = Counter()
    for w in warnings:
        if "PDF 없음" in w:
            kinds["pdf_missing"] += 1
        elif "KV 값 없음" in w:
            kinds["kv_empty"] += 1
        elif "중복 파일" in w:
            kinds["duplicate"] += 1
        else:
            kinds["other"] += 1

    if kinds["pdf_missing"]:
        print(f"[경고] dataset에 PDF 없음: {kinds['pdf_missing']}행 건너뜀")
    if kinds["kv_empty"]:
        print(f"[경고] KV 값 없음: {kinds['kv_empty']}행 건너뜀")
    if kinds["duplicate"]:
        print(f"[경고] 중복 파일: {kinds['duplicate']}건 (마지막 행 유지)")
    other = [w for w in warnings if "PDF 없음" not in w and "KV 값 없음" not in w and "중복 파일" not in w]
    for w in other[:preview]:
        print(f"[경고] {w}")
    extra = len(other) - preview
    if extra > 0:
        print(f"[경고] … 외 {extra}건")


if __name__ == "__main__":
    main()
