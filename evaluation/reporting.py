from __future__ import annotations

from collections import defaultdict
from pathlib import Path
from typing import Callable

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .edit_distance import tokenize_words
from .geometry import format_xywh
from .types import BBox, FileClassMetrics, TextAggregate
from .via import UNVERIFIABLE_GT_VALUE

SEP = "-" * 80


def format_file_class_metrics(
    m: FileClassMetrics, *, kv_scoring: str = "edit_distance"
) -> str:
    n_gt, n_pr = len(m.gt_rects), len(m.pred_rects)
    if kv_scoring == "char_multiset":
        score_line = (
            f"  CER(1-재현율, multiset)={m.cer:.4f} "
            f"(|GT|+|PRED|-2×교집합={m.edit_distance}, 교집합_문자수={m.char_intersection})\n"
        )
    else:
        score_line = f"  CER={m.cer:.4f} (char_dist={m.edit_distance})\n"
    return (
        f"file={m.filename}\n"
        f"  class={m.class_name}\n"
        f"  ref({len(m.ref)}): {m.ref!r}\n"
        f"  hyp({len(m.hyp)}): {m.hyp!r}\n"
        f"{score_line}"
        f"  WER={m.wer:.4f} (word_dist={m.word_edit_distance})\n"
        f"  Word F1={m.word_f1:.4f} (TP={m.word_tp}, FP={m.word_fp}, FN={m.word_fn})\n"
        f"  bbox: GT 원본 {n_gt}개 → 합침 ({format_xywh(m.gt_merged)})\n"
        f"        pred 원본 {n_pr}개 → 합침 ({format_xywh(m.pred_merged)})\n"
        f"  IoU 비교: GT {format_xywh(m.gt_merged)}\n"
        f"            vs pred {format_xywh(m.pred_merged)}\n"
        f"  IoU={m.iou:.4f}\n\n"
    )


def print_file_class_metrics(
    m: FileClassMetrics, *, kv_scoring: str = "edit_distance"
) -> None:
    print(format_file_class_metrics(m, kv_scoring=kv_scoring), end="")


def _fmt_opt(v: float | None, digits: int = 4) -> str:
    if v is None:
        return "n/a"
    return f"{v:.{digits}f}"


def _append_global_summary_lines(
    lines: list[str],
    metrics: list[FileClassMetrics],
    agg: TextAggregate,
    excluded_unverifiable: int = 0,
) -> None:
    if excluded_unverifiable:
        lines.append(
            f"채점 제외(GT '{UNVERIFIABLE_GT_VALUE}'): {excluded_unverifiable}건"
        )
    if metrics:
        lines.append(f"파일·class 항목 수: {len(metrics)}")
        lines.append(
            f"Macro CER: {_fmt_opt(agg.macro_cer)}, "
            f"Micro CER: {_fmt_opt(agg.micro_cer)} "
            f"(문자 편집거리 합 / GT 문자 수)"
        )
        lines.append(
            f"Macro WER: {_fmt_opt(agg.macro_wer)}, "
            f"Micro WER: {_fmt_opt(agg.micro_wer)} "
            f"(단어 편집거리 합 / GT 단어 수)"
        )
        lines.append(
            f"Macro F1: {_fmt_opt(agg.macro_f1)}, "
            f"Micro F1: {_fmt_opt(agg.micro_f1)} "
            f"(단어 단위 TP={agg.total_tp}, FP={agg.total_fp}, FN={agg.total_fn})"
        )
    else:
        lines.append("CER/WER/F1: (항목 없음)")
    if agg.mean_iou is not None:
        lines.append(
            f"평균 IoU (병합 bbox): {agg.mean_iou:.4f} "
            f"(파일·class 항목 수: {len(agg.per_pair_iou)})"
        )
    else:
        lines.append("평균 IoU: (항목 없음)")


def print_summary(
    metrics: list[FileClassMetrics],
    agg: TextAggregate,
    *,
    kv_scoring: str = "edit_distance",
) -> None:
    print(SEP)
    if kv_scoring == "char_multiset":
        print("kv 문자 채점: char_multiset (CER=1-문자 multiset 재현율)")
    else:
        print("kv 문자 채점: edit_distance (Levenshtein CER / WER)")
    if metrics:
        print(f"항목 수: {len(metrics)}")
        print(
            f"Macro CER: {_fmt_opt(agg.macro_cer)}, "
            f"Micro CER: {_fmt_opt(agg.micro_cer)}"
        )
        print(
            f"Macro WER: {_fmt_opt(agg.macro_wer)}, "
            f"Micro WER: {_fmt_opt(agg.micro_wer)}"
        )
        print(
            f"Macro F1: {_fmt_opt(agg.macro_f1)}, "
            f"Micro F1: {_fmt_opt(agg.micro_f1)} "
            f"(TP={agg.total_tp}, FP={agg.total_fp}, FN={agg.total_fn})"
        )
    if agg.mean_iou is not None:
        print(
            f"평균 IoU (병합 bbox): {agg.mean_iou:.4f} "
            f"(파일·class 항목 수: {len(agg.per_pair_iou)})"
        )


def write_detail_txt(
    path: Path, metrics: list[FileClassMetrics], *, kv_scoring: str = "edit_distance"
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        for m in metrics:
            f.write(format_file_class_metrics(m, kv_scoring=kv_scoring))


def _format_rects_list(rects: list[BBox]) -> str:
    """개별 bbox를 한 셀에 넣기 위한 문자열."""
    if not rects:
        return ""
    parts = [format_xywh(b) for b in rects]
    return " | ".join(parts)


def _apply_header_row(ws: Worksheet, headers: list[str], header_font: Font) -> None:
    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).font = header_font


def _autofit_columns(
    ws: Worksheet, headers: list[str], min_w: float, max_w: float
) -> None:
    for col_idx, h in enumerate(headers, start=1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = min(max_w, max(min_w, len(str(h)) + 2))


def write_detail_xlsx(
    path: Path,
    metrics: list[FileClassMetrics],
    common: list[str],
    *,
    kv_scoring: str = "edit_distance",
) -> None:
    """파일별 넓은 표와 파일·class별 긴 표를 저장."""
    path.parent.mkdir(parents=True, exist_ok=True)

    by_key: dict[tuple[str, str], FileClassMetrics] = {
        (m.filename, m.class_name): m for m in metrics
    }
    all_classes = sorted({m.class_name for m in metrics})

    cer_header = "CER(1-재현율)" if kv_scoring == "char_multiset" else "CER"
    dist_header = (
        "|GT|+|PRED|-2×교집합" if kv_scoring == "char_multiset" else "char_edit_dist"
    )

    wb = Workbook()

    ws_wide = wb.active
    ws_wide.title = "파일별"
    header_font = Font(bold=True)
    wide_headers: list[str] = ["파일명"]
    for cls in all_classes:
        wide_headers.extend(
            [
                f"{cls}_Pred",
                f"{cls}_GT",
                f"{cls}_{cer_header}",
                f"{cls}_WER",
                f"{cls}_Word_F1",
                f"{cls}_Score",
                f"{cls}_Pred_bbox",
                f"{cls}_GT_bbox",
                f"{cls}_IoU",
            ]
        )
    ws_wide.append(wide_headers)
    _apply_header_row(ws_wide, wide_headers, header_font)

    for fn in common:
        row: list[str | float] = [fn]
        for cls in all_classes:
            m = by_key.get((fn, cls))
            if m is None:
                row.extend(["", "", "", "", "", "", "", "", ""])
            else:
                row.extend(
                    [
                        m.hyp,
                        m.ref,
                        m.cer,
                        m.wer,
                        m.word_f1,
                        "",  # Score: eval_llm 채점 후 채워짐
                        format_xywh(m.pred_merged),
                        format_xywh(m.gt_merged),
                        m.iou,
                    ]
                )
        ws_wide.append(row)

    ws_wide.freeze_panes = "B2"
    _autofit_columns(ws_wide, wide_headers, min_w=12, max_w=48)

    ws_long = wb.create_sheet("파일·class별")
    long_headers = [
        "파일명",
        "KV class",
        "Pred",
        "GT",
        cer_header,
        "WER",
        "Word_F1",
        "Score",
        dist_header,
        "word_edit_dist",
        "word_TP",
        "word_FP",
        "word_FN",
        "문자_교집합(multiset)",
        "IoU",
        "Pred_bbox_merged",
        "GT_bbox_merged",
        "Pred_bbox_개별",
        "GT_bbox_개별",
    ]
    ws_long.append(long_headers)
    _apply_header_row(ws_long, long_headers, header_font)

    for m in metrics:
        ws_long.append(
            [
                m.filename,
                m.class_name,
                m.hyp,
                m.ref,
                m.cer,
                m.wer,
                m.word_f1,
                "",  # Score: eval_llm 채점 후 채워짐
                m.edit_distance,
                m.word_edit_distance,
                m.word_tp,
                m.word_fp,
                m.word_fn,
                m.char_intersection,
                m.iou,
                format_xywh(m.pred_merged),
                format_xywh(m.gt_merged),
                _format_rects_list(m.pred_rects),
                _format_rects_list(m.gt_rects),
            ]
        )

    ws_long.freeze_panes = "A2"
    _autofit_columns(ws_long, long_headers, min_w=14, max_w=56)

    wb.save(path)


def _group_text_metrics(
    ms: list[FileClassMetrics],
) -> tuple[float, float, float, float, float, float, int]:
    """한 그룹: macro/micro CER·WER·F1, 항목 수."""
    n = len(ms)
    if n == 0:
        return 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0
    macro_cer = sum(m.cer for m in ms) / n
    macro_wer = sum(m.wer for m in ms) / n
    macro_f1 = sum(m.word_f1 for m in ms) / n
    ref_chars = sum(len(m.ref) for m in ms)
    ref_words = sum(len(tokenize_words(m.ref)) for m in ms)
    micro_cer = (
        sum(m.edit_distance for m in ms) / ref_chars if ref_chars > 0 else 0.0
    )
    micro_wer = (
        sum(m.word_edit_distance for m in ms) / ref_words if ref_words > 0 else 0.0
    )
    total_tp = sum(m.word_tp for m in ms)
    total_fp = sum(m.word_fp for m in ms)
    total_fn = sum(m.word_fn for m in ms)
    denom = 2 * total_tp + total_fp + total_fn
    micro_f1 = (2 * total_tp / denom) if denom > 0 else 0.0
    return macro_cer, micro_cer, macro_wer, micro_wer, macro_f1, micro_f1, n


def _aggregate_rows_by_key(
    metrics: list[FileClassMetrics],
    key: Callable[[FileClassMetrics], str],
) -> list[tuple[str, float, float, float, float, float, float, int]]:
    """metrics를 key로 묶어 CER/WER/F1 집계 행을 만든다."""
    grouped: dict[str, list[FileClassMetrics]] = defaultdict(list)
    for m in metrics:
        grouped[key(m)].append(m)
    rows: list[tuple[str, float, float, float, float, float, float, int]] = []
    for k in sorted(grouped.keys()):
        mac_c, mic_c, mac_w, mic_w, mac_f1, mic_f1, n = _group_text_metrics(grouped[k])
        rows.append((k, mac_c, mic_c, mac_w, mic_w, mac_f1, mic_f1, n))
    return rows


def write_summary_txt(
    path: Path,
    metrics: list[FileClassMetrics],
    agg: TextAggregate,
    excluded_unverifiable: int = 0,
    *,
    kv_scoring: str = "edit_distance",
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    lines: list[str] = []

    lines.append("전체 요약")
    lines.append("-" * 40)
    if kv_scoring == "char_multiset":
        lines.append(
            "kv 문자 채점: char_multiset (CER=1-문자 multiset 재현율; WER은 원문 기준)"
        )
    else:
        lines.append("kv 문자 채점: edit_distance (Levenshtein CER / WER)")
    _append_global_summary_lines(lines, metrics, agg, excluded_unverifiable)

    lines.append("")
    lines.append(
        "전체 파일에 대한 class별 집계 "
        "(Macro: 항목 산술평균, Micro: 편집거리 합/GT 길이 합)"
    )
    lines.append("-" * 40)
    class_rows = _aggregate_rows_by_key(metrics, lambda m: m.class_name)
    if class_rows:
        w_class = max(len(r[0]) for r in class_rows)
        lines.append(
            f"{'class':<{w_class}}  {'Macro_CER':>10}  {'Micro_CER':>10}  "
            f"{'Macro_WER':>10}  {'Micro_WER':>10}  {'Macro_F1':>10}  {'Micro_F1':>10}  {'n':>4}"
        )
        for cls, mac_c, mic_c, mac_w, mic_w, mac_f1, mic_f1, n in class_rows:
            lines.append(
                f"{cls:<{w_class}}  {mac_c:10.4f}  {mic_c:10.4f}  "
                f"{mac_w:10.4f}  {mic_w:10.4f}  {mac_f1:10.4f}  {mic_f1:10.4f}  {n:4d}"
            )
    else:
        lines.append("(항목 없음)")

    lines.append("")
    lines.append(
        "전체 class에 대한 파일별 집계 "
        "(Macro: 항목 산술평균, Micro: 편집거리 합/GT 길이 합)"
    )
    lines.append("-" * 40)
    file_rows = _aggregate_rows_by_key(metrics, lambda m: m.filename)
    if file_rows:
        w_file = max(len(r[0]) for r in file_rows)
        lines.append(
            f"{'filename':<{w_file}}  {'Macro_CER':>10}  {'Micro_CER':>10}  "
            f"{'Macro_WER':>10}  {'Micro_WER':>10}  {'Macro_F1':>10}  {'Micro_F1':>10}  {'n_cls':>5}"
        )
        for fn, mac_c, mic_c, mac_w, mic_w, mac_f1, mic_f1, n in file_rows:
            lines.append(
                f"{fn:<{w_file}}  {mac_c:10.4f}  {mic_c:10.4f}  "
                f"{mac_w:10.4f}  {mic_w:10.4f}  {mac_f1:10.4f}  {mic_f1:10.4f}  {n:5d}"
            )
    else:
        lines.append("(항목 없음)")

    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
